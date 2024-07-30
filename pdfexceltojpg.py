import os
from pdf2image import convert_from_path
import openpyxl
import matplotlib.pyplot as plt

def pdf_to_images(pdf_path, output_dir, poppler_path):
    pages = convert_from_path(pdf_path, poppler_path=poppler_path)
    for i, page in enumerate(pages):
        image_path = os.path.join(output_dir, f'page_{i + 1}.png')
        page.save(image_path, 'PNG')
        print(f'Saved {image_path}')

def excel_to_images(excel_path, output_dir):
    workbook = openpyxl.load_workbook(excel_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        fig, ax = plt.subplots()
        ax.axis('tight')
        ax.axis('off')
        table = ax.table(cellText=[[cell.value for cell in row] for row in sheet.iter_rows()],
                         cellLoc='center', loc='center')
        image_path = os.path.join(output_dir, f'{sheet_name}.png')
        plt.savefig(image_path, bbox_inches='tight')
        plt.close(fig)
        print(f'Saved {image_path}')

def main():
    pdf_path = '][.pdf'
    excel_path = '554a07380a0b3ddf1962e930dcd2cdedfurniture.xlsx'
    output_dir = 'output_images'
    poppler_path = r'C:\poppler-24.07.0\Library\bin' 

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    pdf_to_images(pdf_path, output_dir, poppler_path)
    excel_to_images(excel_path, output_dir)

if __name__ == "__main__":
    main()
